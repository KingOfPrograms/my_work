"""
Excel 数据对比统计脚本
======================
加载两份数据，分别筛选→统计，并排对比，含百分比差值。
独立脚本，不影响现有代码。

用法:
    python excel_stats_compare.py

修改下方 CONFIG 区域的参数。
"""

from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ============================================================
# 配置区域
# ============================================================

# 数据 A
FILE_A = "test_data_template.xlsx"
SHEET_A = None                           # None 用第一个 sheet
FILTERS_A = {"需求版本": ["430"], "不纳入统计": ["否"]}
LABEL_A = "430"                          # 在表头中显示的标签

# 数据 B
FILE_B = "test_data_template.xlsx"
SHEET_B = None
FILTERS_B = {"需求版本": ["522"], "不纳入统计": ["否"]}
LABEL_B = "522"

# 共用统计参数
MAIN_DIM = "数据集"
SUB_DIMS = ["一级分类"]
STAT_COL = "process_conclusion"
KEEP_STAT_VALUES = []                    # 空 = 保留全部

OUTPUT_FILE = "统计数据_对比.xlsx"

# ============================================================
# 统计逻辑（复用同一套，不改动）
# ============================================================


def _build_rows(df, group_col, stat_col, group_total_map, keep_values):
    rows = []
    for gval in sorted(df[group_col].dropna().unique()):
        grp_total = group_total_map.get(gval, 0)
        for sv in keep_values:
            cnt = ((df[group_col] == gval) & (df[stat_col] == sv)).sum()
            pct = cnt / grp_total if grp_total else 0
            rows.append({
                "分组": str(gval),
                stat_col: str(sv),
                "占比": round(pct * 100, 2),
                "占比_raw": pct,
                "正确/总数": f"{cnt}/{grp_total}",
                "_total": grp_total,
            })
    return rows


def compute_stats(df, stat_col, main_dim, sub_dims, keep_values):
    total = len(df)
    stats = []

    if not keep_values:
        keep_values = sorted(df[stat_col].dropna().unique().tolist())

    # 整体统计
    overall_rows = []
    for sv in keep_values:
        cnt = (df[stat_col] == sv).sum()
        pct = cnt / total if total else 0
        overall_rows.append({
            "分组": "整体",
            stat_col: str(sv),
            "占比": round(pct * 100, 2),
            "占比_raw": pct,
            "正确/总数": f"{cnt}/{total}",
            "_total": total,
        })
    stats.append(("整体统计", pd.DataFrame(overall_rows)))

    # 按主分组
    if main_dim and main_dim in df.columns:
        m_totals = df.groupby(main_dim).size().to_dict()
        stats.append((
            f"按 [{main_dim}] 统计指标",
            pd.DataFrame(_build_rows(df, main_dim, stat_col, m_totals, keep_values)),
        ))

    # 按子分组
    for sdim in sub_dims:
        if sdim not in df.columns:
            continue
        s_totals = df.groupby(sdim).size().to_dict()
        stats.append((
            f"按 [{sdim}] 统计指标",
            pd.DataFrame(_build_rows(df, sdim, stat_col, s_totals, keep_values)),
        ))

    # 主分组 × 子分组 交叉统计
    if main_dim and sub_dims and main_dim in df.columns:
        for m_val in sorted(df[main_dim].dropna().unique()):
            stats.append((f"--- {main_dim}: {m_val} ---", pd.DataFrame()))
            for sdim in sub_dims:
                if sdim not in df.columns:
                    continue
                subset = df[df[main_dim] == m_val]
                s_totals = subset.groupby(sdim).size().to_dict()
                if not s_totals:
                    continue
                stats.append((
                    f"按 [{sdim}] 统计指标",
                    pd.DataFrame(_build_rows(subset, sdim, stat_col, s_totals, keep_values)),
                ))

    return stats


# ============================================================
# 加载与合并
# ============================================================


def load_and_filter(filepath, sheet, filters):
    """读取 Excel → 筛选 → 返回 DataFrame"""
    sheets = pd.read_excel(filepath, sheet_name=None)
    sname = sheet or list(sheets.keys())[0]
    df = sheets[sname]
    for col, vals in filters.items():
        if col in df.columns and vals:
            # 统一转字符串比较，避免整数/字符串类型不匹配
            vals_str = [str(v) for v in vals]
            df = df[df[col].astype(str).isin(vals_str)]
    return df


def merge_stats(stats_a, stats_b, label_a, label_b, stat_col):
    """将两份统计结果合并为对比表列表"""
    merged = []

    for (title_a, df_a), (title_b, df_b) in zip(stats_a, stats_b):
        # 节标题（空 DF 直接透传）
        if df_a.empty and df_b.empty:
            merged.append((title_a, pd.DataFrame(), label_a, label_b, stat_col))
            continue

        # 按「分组 + stat_col 值」对齐
        key_cols = ["分组", stat_col]
        # 确保 key_cols 存在
        for k in key_cols:
            if k not in df_a.columns:
                df_a[k] = ""
            if k not in df_b.columns:
                df_b[k] = ""

        # rename A/B 列
        a_renamed = df_a.rename(columns={
            "正确/总数": f"{label_a} 正确/总数",
            "占比": f"{label_a} 占比",
            "占比_raw": f"{label_a} 占比_raw",
        })
        b_renamed = df_b.rename(columns={
            "正确/总数": f"{label_b} 正确/总数",
            "占比": f"{label_b} 占比",
            "占比_raw": f"{label_b} 占比_raw",
        })

        # merge
        combined = pd.merge(
            a_renamed[[*key_cols, f"{label_a} 正确/总数", f"{label_a} 占比", f"{label_a} 占比_raw"]],
            b_renamed[[*key_cols, f"{label_b} 正确/总数", f"{label_b} 占比", f"{label_b} 占比_raw"]],
            on=key_cols, how="outer",
        ).fillna("-")

        # 百分比差值（百分点）
        def _parse_pct(val):
            try:
                return float(val)
            except (ValueError, TypeError):
                return None

        pct_a = combined[f"{label_a} 占比"].apply(_parse_pct)
        pct_b = combined[f"{label_b} 占比"].apply(_parse_pct)
        combined["百分比增长"] = [
            round(b - a, 2) if (a is not None) and (b is not None) else "-"
            for a, b in zip(pct_a, pct_b)
        ]
        # 填充 NaN（两边都没数据的情况）
        combined["百分比增长"] = combined["百分比增长"].replace("nan", "-").fillna("-")

        merged.append((title_a, combined, label_a, label_b, stat_col))

    return merged


# ============================================================
# 写入 Excel
# ============================================================


def write_compare_excel(file_a, df_a, merged_tables, output_path):
    """保留原始 sheet，新增「统计对比」sheet"""
    wb = load_workbook(file_a)

    sname = "统计对比"
    if sname in wb.sheetnames:
        del wb[sname]
    ws = wb.create_sheet(sname)

    header_a_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_a_font = Font(bold=True, size=11, color="FFFFFF")
    header_b_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    header_b_font = Font(bold=True, size=11, color="FFFFFF")
    diff_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    section_font = Font(bold=True, size=13, color="1F4E79")
    group_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    group_font = Font(bold=True, size=12, color="375623")
    cell_align = Alignment(horizontal="center", vertical="center")
    pct_fmt = "0.00%"
    diff_pct_fmt = "+0.00%;-0.00%"

    current_row = 1
    for title, cdf, label_a, label_b, stat_col in merged_tables:
        if cdf.empty:
            # 节标题
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
            tc = ws.cell(row=current_row, column=1, value=title)
            tc.font = group_font
            tc.fill = group_fill
            tc.alignment = cell_align
            current_row += 1
            continue

        # 节标题
        ncols = len(cdf.columns)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=ncols)
        tc = ws.cell(row=current_row, column=1, value=title)
        tc.font = section_font
        current_row += 1

        # 表头
        for ci, col_name in enumerate(cdf.columns, 1):
            cel = ws.cell(row=current_row, column=ci, value=str(col_name))
            cel.font = header_a_font if label_a in str(col_name) else (header_b_font if label_b in str(col_name) else Font(bold=True, size=11))
            cel.fill = header_a_fill if label_a in str(col_name) else (header_b_fill if label_b in str(col_name) else (diff_fill if "增长" in str(col_name) else PatternFill()))
            cel.alignment = cell_align
        current_row += 1

        # 数据
        for _, rd in cdf.iterrows():
            for ci, col_name in enumerate(cdf.columns, 1):
                val = rd[col_name]
                if isinstance(val, float) and pd.isna(val):
                    val = ""
                cel = ws.cell(row=current_row, column=ci, value=val if val != "-" else "-")
                cel.alignment = cell_align
                if "占比_raw" in col_name:
                    try:
                        cel.value = float(val) if val not in ("-", "") else ""
                        cel.number_format = pct_fmt
                    except (ValueError, TypeError):
                        pass
                if col_name == "百分比增长":
                    try:
                        cel.value = float(val) / 100 if val not in ("-", "") else ""
                        cel.number_format = diff_pct_fmt
                    except (ValueError, TypeError):
                        pass
            current_row += 1

        current_row += 2

    for ci in range(1, 10):
        ws.column_dimensions[get_column_letter(ci)].width = 22

    wb.save(output_path)


# ============================================================
# 主流程
# ============================================================


def main():
    # 加载
    df_a = load_and_filter(FILE_A, SHEET_A, FILTERS_A)
    df_b = load_and_filter(FILE_B, SHEET_B, FILTERS_B)
    print(f"数据A: {len(df_a)} 条, 数据B: {len(df_b)} 条")

    # 分别统计
    stats_a = compute_stats(df_a, STAT_COL, MAIN_DIM, SUB_DIMS, KEEP_STAT_VALUES)
    stats_b = compute_stats(df_b, STAT_COL, MAIN_DIM, SUB_DIMS, KEEP_STAT_VALUES)

    # 合并
    merged = merge_stats(stats_a, stats_b, LABEL_A, LABEL_B, STAT_COL)

    # 打印预览
    for title, cdf, la, lb, sc in merged:
        if cdf.empty:
            print(f"\n  --- {title} ---")
            continue
        print(f"\n{'='*60}")
        print(f"  {title}")
        print(f"{'='*60}")
        print(cdf.to_string(index=False))

    # 写入 Excel
    write_compare_excel(FILE_A, df_a, merged, OUTPUT_FILE)
    print(f"\n已生成: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
