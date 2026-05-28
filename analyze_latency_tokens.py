#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
时延和Token分析脚本
自动识别输入文件的指标类型（token / stream）和模型版本，
在源文件中追加分析 Sheet，每个模型版本生成一张独立的分析表。

支持两种数据格式：
  1. 并列列格式：model_version_{suffix} 作为列名，匹配同后缀的数据列
  2. 标准格式：单一 model_version 列，按列值分组
"""

import os
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import (
    Font, Alignment, Border, Side, PatternFill
)

# 轮次定义: (key, 显示名)
ROUNDS = [
    ('first', '首轮'),
    ('last',  '最后一轮'),
]

# 指标定义: (key, 显示名, [列名匹配关键词...])
# 按这里的顺序输出各指标组
METRIC_DEFS = [
    ('latency',          '时延(ms)',        ['latency']),
    ('ttft',             'TTFT(ms)',        ['ttft']),
    ('stream_duration',  '流式时长(ms)',     ['stream', 'duration']),
    ('input_tokens',     '输入token',       ['input', 'token']),
    ('output_tokens',    '输出token',       ['output', 'token']),
]

SUB_HEADERS = ["均值", "P50", "P75", "P90", "P95"]

TABLE_GAP = 2

# ---- 样式 ----
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="微软雅黑", size=13, bold=True)
DATA_FONT = Font(name="微软雅黑", size=10)
ROW_LABEL_FONT = Font(name="微软雅黑", size=10, bold=True)


def find_model_groups(df):
    """
    识别文件中的模型分组方式，返回:
      groups: [(title, data_df, col_map_for_group), ...]
      found_metrics: 全局指标列表

    格式1（并列列）: 有 model_version_{suffix} 列，每个后缀一个分组
    格式2（标准）:   有 model_version 列，按列值分组
    """
    all_cols = df.columns.tolist()

    # ---- 格式1: 查找 model_version_{suffix} 列 ----
    mv_suffix_cols = [c for c in all_cols if c.lower().startswith('model_version_')]
    if mv_suffix_cols:
        print("检测到并列列格式（多套模型列）")
        groups = []
        all_found_metrics = []
        for mv_col in mv_suffix_cols:
            suffix = mv_col[len('model_version_'):]  # 提取后缀
            title = str(df[mv_col].dropna().iloc[0]) if len(df[mv_col].dropna()) > 0 else suffix

            # 找匹配该后缀的数据列
            col_map = {}
            found_metric_keys = []
            for round_key, _ in ROUNDS:
                for metric_key, _, patterns in METRIC_DEFS:
                    for col in all_cols:
                        cl = col.lower()
                        if (round_key in cl and all(p in cl for p in patterns)
                                and cl.endswith(suffix.lower())):
                            col_map[(round_key, metric_key)] = col
                            if metric_key not in found_metric_keys:
                                found_metric_keys.append(metric_key)
                            break

            found_metrics = [(k, d) for k, d, _ in METRIC_DEFS if k in found_metric_keys]
            if not found_metrics:
                print(f"  警告: 后缀 '{suffix}' 未找到匹配的数据列，跳过")
                continue

            groups.append((title, df, col_map, found_metrics))
            all_found_metrics = found_metrics  # 取最后一个有效的
            print(f"  {title}: {len(found_metrics)} 个指标组")

        return groups, all_found_metrics

    # ---- 格式2: 标准 model_version 列，按值分组 ----
    mv_col_name = None
    for c in all_cols:
        if c.lower().strip() == 'model_version':
            mv_col_name = c
            break

    if mv_col_name is None:
        print("错误: 未找到 model_version 列（标准格式或 model_version_ 前缀列）")
        sys.exit(1)

    print("检测到标准格式（单一 model_version 列）")
    df[mv_col_name] = df[mv_col_name].astype(str)

    # 全局列检测
    col_map, found_metrics = _detect_metrics_for_df(df)

    groups = []
    for mv_val, sub_df in df.groupby(mv_col_name):
        groups.append((mv_val, sub_df, col_map, found_metrics))
        print(f"  {mv_val}: {len(sub_df)} 条数据")

    return groups, found_metrics


def _detect_metrics_for_df(df):
    """为 DataFrame 检测所有可用的指标列"""
    col_map = {}
    found_metric_keys = []
    for round_key, _ in ROUNDS:
        for metric_key, _, patterns in METRIC_DEFS:
            for col in df.columns:
                cl = col.lower()
                if round_key in cl and all(p in cl for p in patterns):
                    col_map[(round_key, metric_key)] = col
                    if metric_key not in found_metric_keys:
                        found_metric_keys.append(metric_key)
                    break

    found_metrics = [(k, d) for k, d, _ in METRIC_DEFS if k in found_metric_keys]
    if not found_metrics:
        print("错误: 未识别到任何指标列")
        print(f"可用列: {df.columns.tolist()}")
        sys.exit(1)

    return col_map, found_metrics


def get_percentiles(series):
    """返回 [均值, P50, P75, P90, P95]，整数"""
    return [
        int(round(series.mean())),
        int(series.quantile(0.50, method='nearest')),
        int(series.quantile(0.75, method='nearest')),
        int(series.quantile(0.90, method='nearest')),
        int(series.quantile(0.95, method='nearest')),
    ]


def build_table(df, col_map, found_metrics):
    table = {}
    for round_key, round_label in ROUNDS:
        row_data = []
        for metric_key, _ in found_metrics:
            key = (round_key, metric_key)
            if key in col_map:
                row_data.extend(get_percentiles(df[col_map[key]]))
            else:
                row_data.extend([0] * 5)
        table[round_label] = row_data
    return table


def write_table(ws, start_row, title, table_data, found_metrics, total_cols):
    """将一个模型版本的分析表写入工作表，返回下一个可用行号。"""
    col_letters = [chr(ord('A') + i) for i in range(total_cols)]
    merge_range = f"A{start_row}:{col_letters[-1]}{start_row}"

    r = start_row

    # 标题行
    ws.merge_cells(merge_range)
    c = ws.cell(row=r, column=1, value=title)
    c.font = TITLE_FONT
    c.alignment = CENTER_ALIGN
    ws.row_dimensions[r].height = 24
    r += 1

    # 一级表头（指标组）
    for i, (_metric_key, display_name) in enumerate(found_metrics):
        start_col = 2 + i * 5
        end_col = start_col + 4
        ws.merge_cells(start_row=r, start_column=start_col, end_row=r, end_column=end_col)
        ws.cell(row=r, column=start_col, value=display_name)

    for col_idx in range(1, total_cols + 1):
        cell = ws.cell(row=r, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    r += 1

    # 二级表头（均值/P50/P75/P90/P95）
    for col_idx in range(1, total_cols + 1):
        cell = ws.cell(row=r, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    for i in range(len(found_metrics)):
        start_col = 2 + i * 5
        for j, h in enumerate(SUB_HEADERS):
            ws.cell(row=r, column=start_col + j, value=h)
    r += 1

    # 数据行
    for round_label, values in table_data.items():
        ws.cell(row=r, column=1, value=round_label)
        for col_idx in range(1, total_cols + 1):
            cell = ws.cell(row=r, column=col_idx)
            cell.font = DATA_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
        ws.cell(row=r, column=1).font = ROW_LABEL_FONT
        for i, val in enumerate(values):
            ws.cell(row=r, column=2 + i, value=val)
        r += 1

    return r + TABLE_GAP


def main():
    input_file = sys.argv[1] if len(sys.argv) > 1 else "test_stream_data.xlsx"

    if not os.path.exists(input_file):
        print(f"错误: 文件不存在 — {input_file}")
        sys.exit(1)

    print(f"读取: {input_file}")
    df = pd.read_excel(input_file, sheet_name=0)
    print(f"数据行数: {len(df)}")

    groups, found_metrics = find_model_groups(df)
    total_cols = 1 + len(found_metrics) * 5

    print(f"识别到的指标: {', '.join(name for _, name in found_metrics)}")
    print(f"模型版本数: {len(groups)}")

    # 追加写入到源文件
    wb = openpyxl.load_workbook(input_file)

    sheet_name = "时延与Token分析"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(title=sheet_name)

    # 列宽
    col_letters = [chr(ord('A') + i) for i in range(total_cols)]
    ws.column_dimensions['A'].width = 10
    for letter in col_letters[1:]:
        ws.column_dimensions[letter].width = 10

    ws.freeze_panes = "A4"

    # 逐模型版本写入
    current_row = 1
    for title, sub_df, col_map, fm in groups:
        table_data = build_table(sub_df, col_map, fm)
        current_row = write_table(ws, current_row, title, table_data, fm, total_cols)

    wb.save(input_file)
    print(f"完成! 分析表已追加至: {input_file} → Sheet: {sheet_name}")
    print(f"  共 {len(groups)} 张表")


if __name__ == '__main__':
    main()
