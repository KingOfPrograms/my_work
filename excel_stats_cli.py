"""
Excel 统计脚本（命令行）
========================
直接运行即可根据配置生成统计 Excel，无需启动 Streamlit。

用法:
    python excel_stats_cli.py

修改下方 CONFIG 区域的参数即可适配不同统计需求。
"""

from pathlib import Path
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

# ============================================================
# 配置区域 — 修改这里的参数
# ============================================================

INPUT_FILE = "test_data_template.xlsx"   # 输入 Excel 文件路径
OUTPUT_FILE = "统计数据_结果.xlsx"         # 输出 Excel 文件路径
SHEET_NAME = None                         # 指定 sheet 名，None 则使用第一个

# 筛选条件: {列名: [要保留的值]}
FILTERS = {
    "不纳入统计": ["否"],
}

# 主分组维度: None 表示不分组
MAIN_DIM = "数据集"

# 子分组维度: 可多选，留空列表表示无子分组
SUB_DIMS = ["一级分类"]

# 统计指标列
STAT_COL = "process_conclusion"

# 保留的指标值: 空列表表示保留全部，如 ["PASS"] 只显示 PASS 行
KEEP_STAT_VALUES = ["PASS"]

# ============================================================
# 统计逻辑
# ============================================================


def _build_rows(df, group_col, stat_col, group_total_map, keep_values):
    """构建分组统计行"""
    rows = []
    for gval in sorted(df[group_col].dropna().unique()):
        grp_total = group_total_map.get(gval, 0)
        for sv in keep_values:
            cnt = ((df[group_col] == gval) & (df[stat_col] == sv)).sum()
            pct = cnt / grp_total if grp_total else 0
            rows.append({
                "分组": str(gval),
                stat_col: str(sv),
                "占比": round(pct * 100, 2),
                "占比_raw": pct,
                "正确/总数": f"{cnt}/{grp_total}",
                "_total": grp_total,
            })
    return rows


def compute_stats(df, stat_col, main_dim, sub_dims, keep_values):
    """返回 [(title, DataFrame), ...] 列表"""
    total = len(df)
    stats = []

    # 保留的指标值：如果为空则取全部
    if not keep_values:
        keep_values = sorted(df[stat_col].dropna().unique().tolist())

    # 整体统计
    overall_rows = []
    for sv in keep_values:
        cnt = (df[stat_col] == sv).sum()
        pct = cnt / total if total else 0
        overall_rows.append({
            "分组": "整体",
            stat_col: str(sv),
            "占比": round(pct * 100, 2),
            "占比_raw": pct,
            "正确/总数": f"{cnt}/{total}",
            "_total": total,
        })
    stats.append(("整体统计", pd.DataFrame(overall_rows)))

    # 按主分组
    if main_dim and main_dim in df.columns:
        m_totals = df.groupby(main_dim).size().to_dict()
        stats.append((
            f"按 [{main_dim}] 统计指标",
            pd.DataFrame(_build_rows(df, main_dim, stat_col, m_totals, keep_values)),
        ))

    # 按子分组
    for sdim in sub_dims:
        if sdim not in df.columns:
            continue
        s_totals = df.groupby(sdim).size().to_dict()
        stats.append((
            f"按 [{sdim}] 统计指标",
            pd.DataFrame(_build_rows(df, sdim, stat_col, s_totals, keep_values)),
        ))

    # 主分组 × 子分组 交叉统计
    if main_dim and sub_dims and main_dim in df.columns:
        for sdim in sub_dims:
            if sdim not in df.columns:
                continue
            for m_val in sorted(df[main_dim].dropna().unique()):
                subset = df[df[main_dim] == m_val]
                s_totals = subset.groupby(sdim).size().to_dict()
                if not s_totals:
                    continue
                stats.append((
                    f"按 [{main_dim}={m_val}] → [{sdim}] 统计指标",
                    pd.DataFrame(_build_rows(subset, sdim, stat_col, s_totals, keep_values)),
                ))

    return stats


def write_stats_to_excel(orig_wb, combined_df, stats_tables, output_path):
    """将统计表写入 Excel，保留原始 sheet"""
    wb = orig_wb

    stats_sname = "统计分析"
    if stats_sname in wb.sheetnames:
        del wb[stats_sname]
    ws = wb.create_sheet(stats_sname)

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    section_font = Font(bold=True, size=13, color="1F4E79")
    cell_align = Alignment(horizontal="center", vertical="center")
    pct_fmt = "0.00%"

    row = 1
    for title, sdf in stats_tables:
        write_cols = [c for c in sdf.columns if c not in ("占比", "_total")]
        col_map = {"占比_raw": "占比"}
        ncols = len(write_cols)

        # 节标题
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        tc = ws.cell(row=row, column=1, value=title)
        tc.font = section_font
        row += 1

        # 表头
        for ci, col_name in enumerate(write_cols, 1):
            label = col_map.get(col_name, str(col_name))
            c = ws.cell(row=row, column=ci, value=label)
            c.font = header_font
            c.fill = header_fill
            c.alignment = cell_align
        row += 1

        # 数据
        for _, rd in sdf.iterrows():
            for ci, col_name in enumerate(write_cols, 1):
                val = rd[col_name]
                if pd.isna(val):
                    val = ""
                cel = ws.cell(row=row, column=ci, value=val)
                cel.alignment = cell_align
                if col_name == "占比_raw":
                    cel.number_format = pct_fmt
            row += 1

        row += 2

    for ci in range(1, 6):
        ws.column_dimensions[get_column_letter(ci)].width = 22

    wb.save(output_path)


# ============================================================
# 主流程
# ============================================================

def main():
    input_path = Path(INPUT_FILE)
    if not input_path.exists():
        print(f"文件不存在: {input_path}")
        return

    # 读取
    all_sheets = pd.read_excel(input_path, sheet_name=None)
    sname = SHEET_NAME or list(all_sheets.keys())[0]
    df = all_sheets[sname]
    print(f"读取 sheet [{sname}]: {len(df)} 条")

    # 筛选
    for col, vals in FILTERS.items():
        if col in df.columns and vals:
            df = df[df[col].isin(vals)]
    print(f"筛选后: {len(df)} 条")

    if df.empty:
        print("筛选后无数据，退出")
        return

    # 统计
    stats = compute_stats(df, STAT_COL, MAIN_DIM, SUB_DIMS, KEEP_STAT_VALUES)

    # 打印
    for title, sdf in stats:
        print(f"\n{'='*50}")
        print(f"  {title}")
        print(f"{'='*50}")
        display_df = sdf.drop(columns=["占比_raw", "_total"], errors="ignore")
        for _, rd in display_df.iterrows():
            vals = "  |  ".join(str(v) for v in rd.values)
            print(f"  {vals}")

    # 输出 Excel
    orig_wb = load_workbook(input_path)
    write_stats_to_excel(orig_wb, df, stats, OUTPUT_FILE)
    print(f"\n已生成: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
