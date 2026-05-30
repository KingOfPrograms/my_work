"""
分层抽样脚本
============
按条件筛选数据后，根据指定列的值分布（当前比例或自定义比例）抽取用例，输出新 Excel。

用法:
    python sample_cases.py

修改下方 CONFIG 区域的参数。
"""

from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# ============================================================
# 配置区域
# ============================================================

INPUT_FILE = "test_data_template.xlsx"
OUTPUT_FILE = "抽样结果.xlsx"
SHEET_NAME = None                         # None 使用第一个 sheet

# 筛选条件: {列名: [要保留的值]}，多个条件 AND 关系，空字典 = 不筛选
FILTERS = {
    "不纳入统计": ["否"],
}

# 抽样依据列：按此列的值分布进行分层抽样
SAMPLE_COL = "process_conclusion"

# 抽样模式: "current" 保持当前比例 / "custom" 自定义比例
MODE = "current"

# 抽取总数
SAMPLE_SIZE = 10

# 自定义比例（仅 MODE="custom" 时生效）：{值: 比例}，比例之和应为 1.0
# 示例：{"合理": 0.6, "不合理": 0.4}
CUSTOM_RATIOS = {}

# ============================================================
# 抽样逻辑
# ============================================================


def stratified_sample(df, col, total_n, mode="current", custom_ratios=None):
    """按列值分层抽样。

    Args:
        df: 待抽样 DataFrame
        col: 分层依据列
        total_n: 目标总数
        mode: "current" | "custom"
        custom_ratios: dict，仅 custom 模式使用

    Returns:
        抽样后的 DataFrame
    """
    if total_n >= len(df):
        print(f"  目标 {total_n} >= 可用 {len(df)}，返回全部数据")
        return df.copy()

    value_counts = df[col].value_counts()
    total = len(df)
    samples = []

    if mode == "custom" and custom_ratios:
        ratios = custom_ratios
    else:
        # 当前比例
        ratios = {val: cnt / total for val, cnt in value_counts.items()}

    print(f"\n  分层抽样 (总目标: {total_n})")
    print(f"  {'值':<20} {'总数':>6} {'比例':>8} {'目标':>6} {'实际':>6}")
    print(f"  {'-'*50}")

    remaining = total_n
    all_selected = set()

    # 按比例分配名额
    allocations = {}
    for val, ratio in ratios.items():
        target = max(1, round(total_n * ratio)) if ratio > 0 else 0
        allocations[val] = target

    # 调整使总和 = total_n
    diff = total_n - sum(allocations.values())
    if diff != 0:
        # 按比例从大到小调整
        sorted_vals = sorted(allocations, key=lambda v: ratios.get(v, 0), reverse=True)
        for val in sorted_vals:
            if diff == 0:
                break
            if diff > 0:
                allocations[val] += 1
                diff -= 1
            else:
                if allocations[val] > 1:
                    allocations[val] -= 1
                    diff += 1

    for val, target in allocations.items():
        pool = df[df[col] == val]
        available = len(pool)
        actual = min(target, available)

        if actual < target:
            print(f"  [!] {val}: 仅有 {available} 条，不足目标 {target} 条")

        sampled = pool.sample(n=actual, random_state=42)
        samples.append(sampled)
        print(f"  {str(val):<20} {available:>6} {ratios.get(val, 0)*100:>7.1f}% {target:>6} {actual:>6}")

    result = pd.concat(samples, ignore_index=True)
    # 打乱顺序
    result = result.sample(frac=1, random_state=42).reset_index(drop=True)
    print(f"  {'-'*50}")
    print(f"  合计: {len(result)} 条")
    return result


# ============================================================
# 主流程
# ============================================================


def main():
    input_path = Path(INPUT_FILE)
    if not input_path.exists():
        print(f"文件不存在: {input_path}")
        return

    # 读取
    all_sheets = pd.read_excel(input_path, sheet_name=None)
    sname = SHEET_NAME or list(all_sheets.keys())[0]
    df = all_sheets[sname]
    print(f"读取 sheet [{sname}]: {len(df)} 条")

    # 筛选
    for col, vals in FILTERS.items():
        if col in df.columns and vals:
            before = len(df)
            df = df[df[col].astype(str).isin([str(v) for v in vals])]
            print(f"筛选 [{col}] in {vals}: {before} → {len(df)} 条")

    if df.empty:
        print("筛选后无数据")
        return

    # 检查抽样列
    if SAMPLE_COL not in df.columns:
        print(f"抽样列 [{SAMPLE_COL}] 不存在，可用列: {list(df.columns)}")
        return

    # 分层抽样
    result = stratified_sample(df, SAMPLE_COL, SAMPLE_SIZE, MODE, CUSTOM_RATIOS)

    # 输出 Excel
    wb = load_workbook(input_path)
    out_sname = "抽样结果"
    if out_sname in wb.sheetnames:
        del wb[out_sname]
    ws = wb.create_sheet(out_sname)

    # 写表头
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for ci, col_name in enumerate(result.columns, 1):
        c = ws.cell(row=1, column=ci, value=str(col_name))
        c.font = header_font
        c.fill = header_fill

    # 写数据
    for ri, row in enumerate(result.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val if not (isinstance(val, float) and pd.isna(val)) else "")

    wb.save(OUTPUT_FILE)
    print(f"\n已生成: {OUTPUT_FILE}")
    print(f"输出 sheet: [{out_sname}]，共 {len(result)} 条")


if __name__ == "__main__":
    main()
