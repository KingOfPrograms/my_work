"""
Excel 错误主题提取脚本
======================
读取 Excel，从"原因"列的 JSON 中提取所有 valid=false 的 error 信息，
新增"错误主题"列，并生成错误频次统计 sheet。

用法:
    python extract_error_topic.py

修改下方 CONFIG 区域的参数。
"""

import json
import re
from collections import Counter
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ============================================================
# 配置区域
# ============================================================

INPUT_FILE = "test_data_template.xlsx"   # 输入文件路径
OUTPUT_FILE = "错误主题提取结果.xlsx"       # 输出文件路径
SHEET_NAME = None                         # None 使用第一个 sheet

ERROR_COL = "错误"                         # 错误列名
REASON_COL = "原因"                        # 原因列名
NEW_COL = "错误主题"                       # 新增列名

# ============================================================
# 解析逻辑
# ============================================================


def parse_all_errors(reason_str):
    """从原因 JSON 中提取所有 valid=false 的 error 数组元素。

    返回: (拼接字符串, [错误列表])
    """
    empty = ("", [])

    if pd.isna(reason_str) or not isinstance(reason_str, str) or not reason_str.strip():
        return empty

    s = reason_str.strip()

    # 处理 } { 之间缺少逗号的情况
    s = re.sub(r'\}\s*\{', '},{', s)

    try:
        data = json.loads(s)
    except json.JSONDecodeError:
        objs = re.findall(r'\{[^{}]*(?:\{[^{}]*\}[^{}]*)*\}', s)
        if not objs:
            return empty
        try:
            data = [json.loads(obj) for obj in objs]
        except json.JSONDecodeError:
            return empty

    if not data or not isinstance(data, list):
        return empty

    all_errors = []
    for obj in data:
        if not isinstance(obj, dict):
            continue
        if obj.get("valid") is False:
            errors = obj.get("error", [])
            if isinstance(errors, list):
                all_errors.extend(errors)

    joined = "; ".join(str(e) for e in all_errors) if all_errors else ""
    return joined, all_errors


# ============================================================
# 主流程
# ============================================================


def main():
    input_path = Path(INPUT_FILE)
    if not input_path.exists():
        print(f"文件不存在: {input_path}")
        return

    # 读取
    all_sheets = pd.read_excel(input_path, sheet_name=None)
    sname = SHEET_NAME or list(all_sheets.keys())[0]
    df = all_sheets[sname]
    print(f"读取 sheet [{sname}]: {len(df)} 条")

    if ERROR_COL not in df.columns:
        print(f"错误: 未找到列 [{ERROR_COL}]，可用列: {list(df.columns)}")
        return
    if REASON_COL not in df.columns:
        print(f"错误: 未找到列 [{REASON_COL}]，可用列: {list(df.columns)}")
        return

    # 筛选 错误列 == False 的行
    mask = df[ERROR_COL] == False
    target = df[mask]
    target_count = mask.sum()
    print(f"错误列 == False: {target_count} 条")

    # 提取错误主题 + 全局计数器
    error_counter = Counter()
    df[NEW_COL] = ""

    for idx in target.index:
        joined, errors = parse_all_errors(df.at[idx, REASON_COL])
        df.at[idx, NEW_COL] = joined
        for e in errors:
            error_counter[str(e)] += 1
        if joined:
            print(f"  行 {idx}: {joined[:80]}{'...' if len(joined) > 80 else ''}")

    filled = (df[NEW_COL] != "").sum()
    print(f"已提取错误主题: {filled} 条")
    print(f"去重错误主题: {len(error_counter)} 种")

    # --- 输出 Excel ---
    wb = load_workbook(input_path)
    ws = wb[sname]

    # 新增列
    ncols = ws.max_column
    new_col_idx = ncols + 1

    header_fill_red = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    cell_align_center = Alignment(horizontal="center", vertical="center")
    cell_align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # 写列头
    hc = ws.cell(row=1, column=new_col_idx, value=NEW_COL)
    hc.font = header_font_white
    hc.fill = header_fill_red
    hc.alignment = cell_align_center

    # 写数据
    for pd_idx, row_data in df.iterrows():
        excel_row = pd_idx + 2
        val = row_data[NEW_COL]
        if val:
            cell = ws.cell(row=excel_row, column=new_col_idx, value=val)
            cell.alignment = cell_align_left

    ws.column_dimensions[get_column_letter(new_col_idx)].width = 50

    # --- 新增统计 sheet ---
    stats_sname = "错误统计"
    if stats_sname in wb.sheetnames:
        del wb[stats_sname]
    ws2 = wb.create_sheet(stats_sname)

    header_fill_blue = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_bold = Font(bold=True, size=11, color="FFFFFF")

    # 表头
    ws2.cell(row=1, column=1, value="错误主题").font = header_font_bold
    ws2.cell(row=1, column=1).fill = header_fill_blue
    ws2.cell(row=1, column=1).alignment = cell_align_center

    ws2.cell(row=1, column=2, value="出现次数").font = header_font_bold
    ws2.cell(row=1, column=2).fill = header_fill_blue
    ws2.cell(row=1, column=2).alignment = cell_align_center

    # 按出现次数降序排列
    sorted_errors = error_counter.most_common()
    for i, (topic, count) in enumerate(sorted_errors, 2):
        ws2.cell(row=i, column=1, value=topic).alignment = cell_align_left
        c2 = ws2.cell(row=i, column=2, value=count)
        c2.alignment = cell_align_center
        c2.font = Font(bold=True, color="C00000")

    # 合计行
    total_row = len(sorted_errors) + 2
    ws2.cell(row=total_row, column=1, value="合计").font = Font(bold=True)
    ws2.cell(row=total_row, column=1).alignment = cell_align_center
    ws2.cell(row=total_row, column=2, value=sum(error_counter.values())).font = Font(bold=True)
    ws2.cell(row=total_row, column=2).alignment = cell_align_center

    ws2.column_dimensions["A"].width = 55
    ws2.column_dimensions["B"].width = 14

    wb.save(OUTPUT_FILE)
    print(f"\n错误统计（按频次降序）:")
    for topic, count in sorted_errors:
        print(f"  [{count:3d}] {topic}")
    print(f"\n已生成: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
