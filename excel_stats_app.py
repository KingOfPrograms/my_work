"""
Excel 勾选统计工具
==================
三步统计模型：筛选条件 -> 输出指标列（分组维度） -> 统计指标列（计数）。
输出含整体统计和按输出指标分组统计的 Excel。

用法: streamlit run excel_stats_app.py
"""

import io
import streamlit as st
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Excel 数据统计工具", layout="wide")
st.title("Excel 数据统计工具")

# ---------------------------------------------------------------------------
# 1. 加载数据
# ---------------------------------------------------------------------------
DEFAULT_FILE = "test_data_template.xlsx"

uploaded = st.file_uploader("上传 Excel 文件（留空则使用默认文件）", type=["xlsx", "xls"])
source = uploaded if uploaded else DEFAULT_FILE

try:
    if isinstance(source, str):
        all_sheets = pd.read_excel(source, sheet_name=None)
        import openpyxl as _xl
        _orig_wb = _xl.load_workbook(source)
    else:
        all_sheets = pd.read_excel(source, sheet_name=None)
        _orig_wb = None
except Exception as e:
    st.error(f"文件读取失败: {e}")
    st.stop()

sheet_names = list(all_sheets.keys())
tab_labels = sheet_names if len(sheet_names) > 1 else sheet_names
tabs = st.tabs(tab_labels)

# ---------------------------------------------------------------------------
# 2. 数据表格（带勾选框）
# ---------------------------------------------------------------------------
selected_indices = {}

for idx, sname in enumerate(sheet_names):
    df = all_sheets[sname].copy()
    with tabs[idx]:
        st.subheader(sname)

        df_display = df.copy()
        df_display.insert(0, "选择", True)

        edited = st.data_editor(
            df_display,
            use_container_width=True,
            hide_index=True,
            column_config={"选择": st.column_config.CheckboxColumn(width="small")},
            num_rows="dynamic",
            key=f"editor_{sname}",
        )

        selected_mask = edited["选择"].values if "选择" in edited.columns else []
        selected_idx = (
            df.index[selected_mask].tolist()
            if len(selected_mask) == len(df)
            else df.index.tolist()
        )
        selected_indices[sname] = selected_idx
        st.caption(f"已选 {len(selected_idx)} / {len(df)} 条")

# ---------------------------------------------------------------------------
# 3. 统计配置（三步模型）
# ---------------------------------------------------------------------------
st.divider()
st.subheader("统计配置")

# 收集选中数据
all_data = []
for sname in sheet_names:
    df = all_sheets[sname]
    idx = selected_indices.get(sname, [])
    selected = df.loc[df.index.isin(idx)].copy()
    all_data.append(selected)

combined = pd.concat(all_data, ignore_index=True) if all_data else pd.DataFrame()
available_cols = list(combined.columns)

st.info(f"当前选中数据: {len(combined)} 条")

# --- Step 1: 筛选条件 ---
st.markdown("### Step 1: 筛选条件")
st.caption("根据某列的值筛选参与统计的用例，多个条件之间为 AND 关系")

if "filters" not in st.session_state:
    st.session_state.filters = []

def add_filter():
    st.session_state.filters.append({
        "col": available_cols[0] if available_cols else None,
        "values": [],
    })

def remove_filter(i):
    st.session_state.filters.pop(i)

st.button("+ 添加筛选条件", on_click=add_filter)

filtered_df = combined.copy()
for i, f in enumerate(st.session_state.filters):
    fc1, fc2, fc3 = st.columns([3, 4, 1])
    with fc1:
        col_list = [f.get("col")] if f.get("col") else []
        for c in available_cols:
            if c not in col_list:
                col_list.append(c)
        f["col"] = st.selectbox(
            "筛选列",
            available_cols,
            key=f"flt_col_{i}",
            index=available_cols.index(f["col"]) if f.get("col") in available_cols else 0,
        )
    with fc2:
        if f.get("col") and f["col"] in filtered_df.columns:
            col_vals = sorted(filtered_df[f["col"]].dropna().unique().tolist())
        else:
            col_vals = []
        f["values"] = st.multiselect("等于（可多选）", col_vals, key=f"flt_vals_{i}")
    with fc3:
        st.button("X 删除", key=f"flt_del_{i}", on_click=remove_filter, args=(i,))

# 应用筛选
for f in st.session_state.filters:
    vals = f.get("values", [])
    col = f.get("col")
    if vals and col and col in filtered_df.columns:
        filtered_df = filtered_df[filtered_df[col].isin(vals)]

st.caption(f"筛选后: {len(filtered_df)} 条（原始 {len(combined)} 条）")

# --- Step 2: 输出指标列 ---
st.markdown("### Step 2: 输出指标列（可多选）")
st.caption("按这些列的值分组，生成各自的统计表（如按 [数据集] 和一 [一级分类] 分别统计）")

output_dims = st.multiselect(
    "选择输出指标列（可多选，空则仅输出整体）",
    options=available_cols,
    key="output_dims",
)

# --- Step 3: 统计指标列 ---
st.markdown("### Step 3: 统计指标列")
st.caption("统计此列各值的数量和占比（如 process_conclusion 中各值的分布）")

stat_col = st.selectbox(
    "选择统计指标列",
    options=available_cols,
    key="stat_col",
)

# ---------------------------------------------------------------------------
# 4. 执行统计 & 输出 Excel
# ---------------------------------------------------------------------------
if st.button("执行统计并生成 Excel", type="primary"):
    if filtered_df.empty:
        st.warning("筛选后无数据")
        st.stop()

    if not stat_col or stat_col not in filtered_df.columns:
        st.warning("请选择统计指标列")
        st.stop()

    total = len(filtered_df)
    stats_tables = []  # (title, DataFrame) pairs

    # === Part 1: 整体统计 ===
    overall_rows = []
    for sv in sorted(filtered_df[stat_col].dropna().unique()):
        cnt = (filtered_df[stat_col] == sv).sum()
        overall_rows.append({
            "分组": "整体",
            stat_col: str(sv),
            "占比": f"{cnt / total * 100:.1f}%",
            "正确/总数": f"{cnt}/{total}",
        })
    overall_df = pd.DataFrame(overall_rows)
    stats_tables.append(("整体统计", overall_df))

    # === Part 2: 按每个输出指标分组统计 ===
    for dim in output_dims:
        if dim not in filtered_df.columns:
            continue

        pivoted = (
            filtered_df.groupby([dim, stat_col])
            .size()
            .reset_index(name="数量")
        )

        dim_totals = filtered_df.groupby(dim).size().to_dict()

        dim_rows = []
        for dim_val in sorted(filtered_df[dim].dropna().unique()):
            grp_total = dim_totals.get(dim_val, 0)
            for sv in sorted(filtered_df[stat_col].dropna().unique()):
                cnt = pivoted[(pivoted[dim] == dim_val) & (pivoted[stat_col] == sv)]["数量"].sum()
                dim_rows.append({
                    "分组": str(dim_val),
                    stat_col: str(sv),
                    "占比": f"{cnt / grp_total * 100:.1f}%" if grp_total > 0 else "0.0%",
                    "正确/总数": f"{cnt}/{grp_total}",
                })

        dim_df = pd.DataFrame(dim_rows)
        stats_tables.append((f"按 [{dim}] 统计指标", dim_df))

    # --- 页面预览 ---
    for title, sdf in stats_tables:
        st.markdown(f"**{title}**")
        st.dataframe(sdf, use_container_width=True, hide_index=True)

    # --- 输出 Excel ---
    output = io.BytesIO()

    if _orig_wb is not None:
        wb = _orig_wb
    else:
        wb = _xl.Workbook()
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
            ws_default = wb["Sheet"]
            for ci, col_name in enumerate(combined.columns, 1):
                ws_default.cell(row=1, column=ci, value=col_name)
            for ri, row in combined.itertuples(index=False):
                for ci, val in enumerate(row, 1):
                    ws_default.cell(row=ri + 2, column=ci, value=val)

    stats_sname = "统计分析"
    if stats_sname in wb.sheetnames:
        del wb[stats_sname]
    ws_stats = wb.create_sheet(stats_sname)

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    section_font = Font(bold=True, size=13, color="1F4E79")
    cell_align = Alignment(horizontal="center", vertical="center")

    current_row = 1
    for title, sdf in stats_tables:
        ncols = len(sdf.columns)
        # 节标题
        ws_stats.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=ncols)
        title_cell = ws_stats.cell(row=current_row, column=1, value=title)
        title_cell.font = section_font
        current_row += 1

        # 表头
        for ci, col_name in enumerate(sdf.columns, 1):
            cell = ws_stats.cell(row=current_row, column=ci, value=str(col_name))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = cell_align
        current_row += 1

        # 数据行
        for _, row_data in sdf.iterrows():
            for ci, val in enumerate(row_data, 1):
                cel = ws_stats.cell(row=current_row, column=ci, value=val if not pd.isna(val) else "")
                cel.alignment = cell_align
            current_row += 1

        current_row += 2

    for ci in range(1, 6):
        ws_stats.column_dimensions[get_column_letter(ci)].width = 22

    wb.save(output)
    output.seek(0)

    st.success("统计完成，点击下方按钮下载")
    st.download_button(
        label="下载统计 Excel",
        data=output,
        file_name="统计数据_结果.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )
